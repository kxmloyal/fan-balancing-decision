# Excel导出器模块
import os
from datetime import datetime

class ExcelExporter:
    def __init__(self, report_exporter):
        """
        Excel导出器
        
        Args:
            report_exporter: 报告导出器实例
        """
        self.report_exporter = report_exporter
        self.output_folder = getattr(report_exporter, 'output_folder', 'outputs')
        self.excel_available = False
        
        # 尝试导入openpyxl库
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            self.excel_available = True
            self.Workbook = Workbook
            self.Font = Font
            self.Alignment = Alignment
            self.PatternFill = PatternFill
            self.Border = Border
            self.Side = Side
            self.get_column_letter = get_column_letter
        except ImportError:
            print("openpyxl库不可用，Excel导出功能将不可用")
            print("请安装openpyxl库：pip install openpyxl")
    
    def export(self, session_data, output_filename=None, task_id=None):
        """
        从会话数据导出Excel格式报告
        
        Args:
            session_data: 会话数据，包含分析结果
            output_filename: 输出文件名
            task_id: 任务ID（可选）
            
        Returns:
            str: Excel文件路径
            
        Raises:
            Exception: 当openpyxl库不可用或导出失败时抛出异常
        """
        # 检查openpyxl是否可用
        if not self.excel_available:
            raise Exception("Excel导出功能依赖的openpyxl库不可用，请安装openpyxl库：pip install openpyxl")
        
        try:
            # 如果提供了任务ID，更新任务状态
            if task_id:
                self.report_exporter.task_manager.update_task_status(
                    task_id, 'in_progress', progress=20, message='开始创建Excel工作簿'
                )
            
            # 确保output_folder属性存在
            os.makedirs(self.output_folder, exist_ok=True)
            
            # 生成默认输出文件名
            if not output_filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_filename = f"report_{timestamp}.xlsx"
            
            # 确保文件名以.xlsx结尾
            if not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'
            
            # 构建输出路径
            output_path = os.path.join(self.output_folder, output_filename)
            
            # 更新任务进度
            if task_id:
                self.report_exporter.task_manager.update_task_status(
                    task_id, 'in_progress', progress=30, message='创建Excel工作簿'
                )
            
            # 创建Excel工作簿
            wb = self.Workbook()
            
            # 获取默认工作表
            ws_summary = wb.active
            ws_summary.title = '分析摘要'
            
            # 填充报告信息
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            fan_model = session_data.get('fan_model', '未知')
            
            # 获取最优转速
            best_speed = "未知"
            if 'evaluation_report' in session_data and 'best_speeds' in session_data['evaluation_report']:
                best_speed = session_data['evaluation_report']['best_speeds'][0] if session_data['evaluation_report']['best_speeds'] else "未知"
            
            # 更新任务进度
            if task_id:
                self.report_exporter.task_manager.update_task_status(
                    task_id, 'in_progress', progress=40, message='添加报告标题和基本信息'
                )
            
            # 添加报告标题
            ws_summary['A1'] = '设备不平衡量分析报告'
            ws_summary['A1'].font = self.Font(bold=True, size=16)
            ws_summary['A1'].alignment = self.Alignment(horizontal='center')
            ws_summary.merge_cells('A1:D1')
            
            # 添加扇叶型号
            ws_summary['A2'] = f'扇叶型号: {fan_model}'
            ws_summary['A2'].font = self.Font(bold=True)
            ws_summary.merge_cells('A2:D2')
            
            # 添加报告信息
            ws_summary['A3'] = f'报告生成时间: {timestamp}'
            ws_summary['A4'] = '报告类型: Excel格式分析报告'
            
            # 添加分析摘要
            ws_summary['A6'] = '分析摘要'
            ws_summary['A6'].font = self.Font(bold=True, underline='single')
            ws_summary.merge_cells('A6:D6')
            
            ws_summary['A7'] = '通过对设备在不同转速下的不平衡量数据进行统计分析，得到以下关键结论：'
            ws_summary.merge_cells('A7:D7')
            
            ws_summary['A8'] = f'推荐最优运行转速：{best_speed}'
            ws_summary['A8'].font = self.Font(bold=True)
            ws_summary.merge_cells('A8:D8')
            
            ws_summary['A9'] = '该转速点是基于IQR（四分位距）和变异系数综合评估确定的，这两个指标反映了数据的离散程度，数值越小表示设备运行越稳定。'
            ws_summary.merge_cells('A9:D9')
            
            # 更新任务进度
            if task_id:
                self.report_exporter.task_manager.update_task_status(
                    task_id, 'in_progress', progress=50, message='添加最优转速选择方法'
                )
            
            # 添加最优转速选择方法
            ws_summary['A11'] = '最优转速选择方法（综合评估）'
            ws_summary['A11'].font = self.Font(bold=True, underline='single')
            ws_summary.merge_cells('A11:D11')
            
            method_steps = [
                '采用三级评估模型确定最优转速：',
                '1. 指标归一化处理：对每个面(P1/P2/ST)分别计算IQR和变异系数(CV)，并进行归一化处理：得分 = 1 / (1 + 指标值)',
                '2. 面内综合得分计算：对每个面的IQR得分和CV得分进行加权综合：面得分 = 0.5 × IQR得分 + 0.5 × CV得分',
                '3. 面间综合总得分计算：根据不同面的重要性进行加权综合：',
                '   - P1面权重：40%',
                '   - P2面权重：40%',
                '   - ST面权重：20%',
                '   - 总得分 = 0.4 × P1得分 + 0.4 × P2得分 + 0.2 × ST得分',
                '4. 最优转速选择：根据总得分排序，得分最高的转速为最优转速'
            ]
            
            for i, step in enumerate(method_steps, 12):
                ws_summary[f'A{i}'] = step
                ws_summary.merge_cells(f'A{i}:D{i}')
            
            # 更新任务进度
            if task_id:
                self.report_exporter.task_manager.update_task_status(
                    task_id, 'in_progress', progress=60, message='添加优化建议'
                )
            
            # 添加优化建议
            ws_summary['A22'] = '优化建议'
            ws_summary['A22'].font = self.Font(bold=True, underline='single')
            ws_summary.merge_cells('A22:D22')
            
            recommendations = [
                '1. 首选推荐转速：建议优先选用推荐的最优运行转速，该转速下设备表现出最佳的运行稳定性',
                '2. 次优转速选择：如果最优转速因工艺限制无法使用，可参考统计表格中其他IQR和CV值较小的转速点',
                '3. 定期监测：建议在选定转速下建立长期监测机制，持续跟踪设备运行状态',
                '4. 数据质量提升：为进一步提高分析准确性，建议增加每组转速下的测量样本数量',
                '5. 多维度评估：除不平衡量外，还可结合温度、振动等其他关键指标进行综合评估'
            ]
            
            for i, recommendation in enumerate(recommendations, 23):
                ws_summary[f'A{i}'] = recommendation
                ws_summary.merge_cells(f'A{i}:D{i}')
            
            # 更新任务进度
            if task_id:
                self.report_exporter.task_manager.update_task_status(
                    task_id, 'in_progress', progress=70, message='添加统计分析结果工作表'
                )
            
            # 添加统计分析结果工作表
            ws_stats = wb.create_sheet(title='统计分析结果')
            
            # 添加表头
            ws_stats['A1'] = '转速'
            ws_stats['B1'] = 'P1面-IQR'
            ws_stats['C1'] = 'P1面-CV'
            ws_stats['D1'] = 'P2面-IQR'
            ws_stats['E1'] = 'P2面-CV'
            ws_stats['F1'] = 'ST面-IQR'
            ws_stats['G1'] = 'ST面-CV'
            ws_stats['H1'] = '综合得分'
            ws_stats['I1'] = '评价'
            
            # 设置表头样式
            header_font = self.Font(bold=True, color='FFFFFF')
            header_fill = self.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_border = self.Border(left=self.Side(style='thin'), right=self.Side(style='thin'), top=self.Side(style='thin'), bottom=self.Side(style='thin'))
            
            for col in range(1, 10):
                cell = ws_stats[self.get_column_letter(col) + '1']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = self.Alignment(horizontal='center')
                cell.border = header_border
            
            # 添加测试数据（实际项目中应从session_data获取）
            test_data = [
                ['2500rpm', 2.2, 0.15, 2.5, 0.18, 1.8, 0.12, 0.95, '最优'],
                ['3000rpm', 3.5, 0.22, 3.8, 0.25, 3.0, 0.18, 0.85, '良好'],
                ['3500rpm', 4.8, 0.28, 5.1, 0.31, 4.2, 0.24, 0.75, '一般'],
                ['4000rpm', 6.2, 0.35, 6.5, 0.38, 5.5, 0.30, 0.65, '较差']
            ]
            
            for i, row in enumerate(test_data, 2):
                for j, value in enumerate(row, 1):
                    ws_stats[self.get_column_letter(j) + str(i)] = value
                    
                    # 设置单元格样式
                    cell = ws_stats[self.get_column_letter(j) + str(i)]
                    cell.border = self.Border(left=self.Side(style='thin'), right=self.Side(style='thin'), top=self.Side(style='thin'), bottom=self.Side(style='thin'))
                    
                    # 高亮最优转速行
                    if value == '最优':
                        for k in range(1, 10):
                            highlight_cell = ws_stats[self.get_column_letter(k) + str(i)]
                            highlight_cell.fill = self.PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
            
            # 更新任务进度
            if task_id:
                self.report_exporter.task_manager.update_task_status(
                    task_id, 'in_progress', progress=80, message='添加统计指标说明工作表'
                )
            
            # 添加统计指标说明工作表
            ws_indicators = wb.create_sheet(title='统计指标说明')
            
            ws_indicators['A1'] = '统计指标说明'
            ws_indicators['A1'].font = self.Font(bold=True, size=14)
            ws_indicators.merge_cells('A1:B1')
            
            indicators = [
                ['平均值', '反映数据的集中趋势'],
                ['中位数', '不受极值影响的中心位置度量'],
                ['标准偏差', '衡量数据的离散程度'],
                ['最小值', '数据中的最小值'],
                ['最大值', '数据中的最大值'],
                ['IQR（四分位距）', '衡量中间50%数据的离散程度，比标准偏差更稳健'],
                ['变异系数(CV)', '标准偏差与平均值的比值，消除了量纲影响，更适合比较不同平均水平的数据波动性']
            ]
            
            for i, (name, desc) in enumerate(indicators, 2):
                ws_indicators['A' + str(i)] = name
                ws_indicators['B' + str(i)] = desc
                ws_indicators['A' + str(i)].font = self.Font(bold=True)
            
            # 添加注意事项工作表
            ws_notes = wb.create_sheet(title='注意事项')
            
            ws_notes['A1'] = '注意事项'
            ws_notes['A1'].font = self.Font(bold=True, size=14)
            ws_notes.merge_cells('A1:B1')
            
            notes = [
                'IQR（四分位距）和变异系数反映了数据的离散程度，数值越小表示数据越稳定',
                '建议关注这些指标较小的转速点，这些点通常代表设备运行较稳定的状态',
                '如需进一步分析，请结合设备的实际运行情况进行综合判断',
                '本报告提供的最优转速建议仅供参考，实际应用中还需考虑工艺要求和其他工程因素',
                '报告中的图表和数据可下载保存，供后续分析和汇报使用'
            ]
            
            for i, note in enumerate(notes, 2):
                ws_notes['A' + str(i)] = note
                ws_notes.merge_cells('A' + str(i) + ':B' + str(i))
            
            # 调整列宽
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = self.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # 更新任务进度
            if task_id:
                self.report_exporter.task_manager.update_task_status(
                    task_id, 'in_progress', progress=90, message='保存Excel文件'
                )
            
            # 保存Excel文件
            wb.save(output_path)
            
            # 添加到导出历史
            export_info = {
                'type': 'excel',
                'filename': os.path.basename(output_path),
                'path': output_path,
                'fan_model': session_data.get('fan_model', '未知')
            }
            self.report_exporter.history_manager.add_record(export_info)
            
            # 更新任务状态为完成
            if task_id:
                self.report_exporter.task_manager.update_task_status(
                    task_id, 'completed', progress=100, message='Excel报告导出完成', result=output_path
                )
            
            return output_path
        except Exception as e:
            # 更新任务状态为失败
            if task_id:
                self.report_exporter.task_manager.update_task_status(
                    task_id, 'failed', progress=0, message='Excel报告导出失败', error=str(e)
                )
            print(f"导出Excel报告失败: {str(e)}")
            raise
